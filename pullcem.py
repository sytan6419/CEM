from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from datetime import *
from PIL import Image
import os
import datetime
import openpyxl
from openpyxl.chart import BarChart, Series, Reference
import time

def pullcmrc():
    list = []
    # cmrc date format = month/date/year
    date = datetime.datetime.now()

    if date.hour < 16:
        print 'It is more accurate if you pull the data after 4pm'
        date_from = '%d/%d/%d'%(date.month,date.day-1,date.year)
        date_to = '%d/%d/%d'%(date.month,date.day-2,date.year)
    else:
        date_from = '%d/%d/%d'%(date.month,date.day-1,date.year)
        date_to = '%d/%d/%d'%(date.month,date.day-1,date.year)

    print 'we are looking for date %s to %s'%(date_from,date_to)
    try:
        # create a textfile for variable input
        f = open('tester.txt','r')
        info = f.read()
        f.close()
    except IOError:
        print 'Please create pidtocheck.txt file on the same directory'
        exit(0)
    except ValueError:
        print 'PID not found'
        exit(0)

    station_prefix = ['C6K','MTP','TCM','C7K','SPA']
    for tester_read in info.split():
        new_test = ''
        if 'pcba' in tester_read:
            for _ in station_prefix:
                try:
                    abc = int(tester_read[-2:])
                    new_test += ', SPEPVM'+_+'2C'+tester_read[-2:]
                except ValueError:
                    new_test += ', SPEPVM'+_+'2C'+tester_read[-1:]
        else:
            new_test = tester_read

        print new_test
        
        for retry in range(1,4):
            try:
                browser = webdriver.Chrome()
                # path2phantom = os.getcwd()+"\phantomjs.exe"
                # browser = webdriver.PhantomJS(executable_path=path2phantom)
                browser.get('http://cpp.cisco.com/CMRC/rpEditorMan.aspx?ReportName=cmrcUUT_History') # change your path at here
                assert 'cmrcLogin' in browser.title
                browser.find_element_by_name('tbxUserID').clear()
                browser.find_element_by_name('tbxUserID').send_keys('') #put your own username
                browser.find_element_by_name('txtPassword').send_keys('') #put your own password 
                browser.find_element_by_name('BtnLogin').click()
                        
                assert 'rpEditorMan' in browser.title
                print 'Logged In'
                browser.find_element_by_name('ctrlItemSelection:txtSerialNumber').send_keys('sal%')
                browser.find_element_by_name('ctrlItemSelection:txtUUTType').send_keys('%')
                browser.find_element_by_name('ctrlItemSelection:txtTestArea').send_keys('FPCB%, PCB%')
                browser.find_element_by_name('ctrlItemSelection:txtCM').send_keys('solpen')
                browser.find_element_by_name('ctrlItemSelection:txtMachine').send_keys(new_test)
                browser.find_element_by_name('Ctrldateselection:cpDateFrom').clear()
                browser.find_element_by_name('Ctrldateselection:cpDateTo').clear()
                browser.find_element_by_name('Ctrldateselection:cpDateFrom').send_keys(date_from)
                browser.find_element_by_name('Ctrldateselection:cpDateTo').send_keys(date_to)
                browser.find_element_by_xpath('//img[contains(@src,"Images/btnSubmit.gif")]').click()

                assert 'rpUUT_HistoryData' in browser.title
                print 'done loading'
                
    ##            # please enable the auto save all file in ur firefox setting if the auto download doesn't work
    ##            browser.find_element_by_id('CtrlGridColDisplay1_btnExportToExcel').click()
    ##            print 'exporting data to excel file ....'
    ##            print 'hold on, we are writing data to info.txt'
    ##
                # writing info to txt file                
                with open("{}.txt".format(tester_read), "w") as text_file:
                    text_file.write(browser.find_element_by_xpath("html").text)
                print 'done writing, will read it soon!'

                print 'logging out'
                browser.find_element_by_link_text("Logout").click();
                print 'will sleep here for 10s to wait download done'
                time.sleep(10)
                browser.quit()

            except AssertionError:
                print 'assertionerror'
                print 'Can\'t get expected output, retrying for %s times'%(retry)
                continue
            except Exception,e:
                print str(e)
                print 'Did u mistakenly close on the web browser?'
                print 'Nvm, we are retrying for %s times'%(retry)
                continue
            break
        else:
            print 'enough of trying ... we are going to terminate this!!!'
            break

            return

def AnalyzeQlik():

    g = open('tester.txt','r')
    g = g.read()

    testers = [t.upper()for t in g.splitlines()]


    f = open('stat.txt','r')
    f = f.read()

    abc = [ _ for _ in f.splitlines() if any(tester in _.split()[0] for tester in testers)]

    fail_list = []
    successlist = []

    abc = [ _ for _ in f.splitlines() if any(tester in _.split()[0] for tester in testers)]

    for _ in abc:
        if '0.00' in _.split()[1] or 'CCPM' in _.split()[0]:
            fail_list.append(_.split()[0])
        elif any(tester == _.split()[0] for tester in testers):
            successlist.append(_.split()[:2])

    fail = [ _ for _ in f.splitlines() if any(fail in _.split()[0] for fail in fail_list)]

    aaa = 0
    new_list = []

    for _ in testers:
        for u in fail:
            if _ in u.split('_'):
                aaa += float(u.split()[1].replace(',',''))
        if (aaa):
            successlist.append(['{}\t\t{}'.format(_.split('_')[0],aaa)])
            aaa = 0
        

    with open('data.txt','w+') as w:
        for i in successlist:
            abc = '\t\t'.join(i)+'\n'
            w.write(abc)
    w.close()
    print 'done'
    print 'total result %d of %d' %(len(successlist),len(testers))
    
    return

 
def AnalyzeCMRC():

    h = open('data.txt','r')
    h = h.read()

    data_list = []

    for _ in h.splitlines():
        data_list.append(_)

    f = open('tester.txt','r')
    f = f.read()

    tester = []
    for _ in f.split():
        tester.append(_.upper())

    new_list = []

    for test in tester:
        try:
            g = open('{}.txt'.format(test),'r')
            g = g.read()
            keyword = ['Aborted','Passed','Failed']
            count = 0
            for i in g.splitlines():
                if any(x for x in keyword if x in i):
                    count += 1
            print count
            print 'Done for {}.txt'.format(test)
            for www in data_list:
                if test == www.split()[0]:
                    kkk = www +'\t' + str(count)+'\n'
                    new_list.append(kkk)
        except IOError:
            print 'No tester file found'
    print 'Done for all'

    print new_list

    hh = open('final.txt','w')

    for _ in new_list:
        hh.write(_)
        
    print 'done'
    hh.close()
    
    return
    

def ExportToExcel():

    h = open('final.txt','r')
    h = h.read()

    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.cell(column=1,row=1,value='Server')
    sheet1.cell(column=2,row=1,value='Consumption')
    sheet1.cell(column=3,row=1,value='Output')
    sheet1.cell(column=4,row=1,value='Average')
    sername = [sername.split()[0] for sername in h.splitlines()]
    consump = [float(consump.split()[1].replace(',','')) for consump in h.splitlines()]
    output = [int(output.split()[2]) for output in h.splitlines()]

    for row in range(len(sername)):
        _ = sheet1.cell(column=1, row=row+2, value="%s" %sername[row])
        _ = sheet1.cell(column=2, row=row+2, value=consump[row])
        _ = sheet1.cell(column=3, row=row+2, value=output[row])
        _ = sheet1.cell(column=4, row=row+2, value="=B%d/C%d" %(row+2,row+2))

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Server vs Consumption"
    chart1.y_axis.title = 'Consumption'
    chart1.x_axis.title = 'Server Name'

    data = Reference(sheet1, min_col=2, min_row=1, max_row=len(sername)+1, max_col=3)
    cats = Reference(sheet1, min_col=1, min_row=2, max_row=len(sername)+1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    sheet1.add_chart(chart1, "I1")

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Server vs Consumption"
    chart1.y_axis.title = 'Consumption'
    chart1.x_axis.title = 'Server Name'

    data = Reference(sheet1, min_col=4, min_row=1, max_row=len(sername)+1, max_col=4)
    cats = Reference(sheet1, min_col=1, min_row=2, max_row=len(sername)+1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    sheet1.add_chart(chart1, "I20")
    global name
    name = "EnergyConsumption_{}.xlsx".format(datetime.datetime.now().date())
    book.save(name)

    return


def main():
    pullcmrc()
    AnalyzeQlik()
    AnalyzeCMRC()
    ExportToExcel()
    print 'Completed for all!!!'
    os.startfile(name)
    return

main()
